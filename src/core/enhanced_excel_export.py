"""
Enhanced Excel Export Module with Formatting and Charts

Professional Excel generation with:
- Advanced formatting (colors, fonts, borders)
- Formulas and calculations
- Charts and graphs
- Multiple sheets
- Data validation
- Conditional formatting
- Freeze panes, filters, and more
"""

from typing import List, Dict, Any, Optional, Union, Tuple
from datetime import datetime, date
from decimal import Decimal
from pathlib import Path
import logging

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
    from openpyxl.chart import (
        BarChart, LineChart, PieChart, AreaChart,
        Reference, Series
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("openpyxl not available. Enhanced Excel export disabled.")

logger = logging.getLogger('dms.excel')


class ExcelStyle:
    """Predefined styles for Excel cells."""

    # Colors
    COLOR_HEADER = "0D6EFD"  # Blue
    COLOR_ACCENT = "198754"  # Green
    COLOR_WARNING = "FFC107"  # Yellow
    COLOR_DANGER = "DC3545"  # Red
    COLOR_LIGHT = "F8F9FA"  # Light gray
    COLOR_DARK = "212529"  # Dark gray

    # Fonts
    FONT_HEADER = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
    FONT_TITLE = Font(name='Calibri', size=16, bold=True, color="000000")
    FONT_BODY = Font(name='Calibri', size=11, color="000000")
    FONT_BOLD = Font(name='Calibri', size=11, bold=True, color="000000")
    FONT_ITALIC = Font(name='Calibri', size=11, italic=True, color="666666")

    # Alignments
    ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
    ALIGN_LEFT = Alignment(horizontal='left', vertical='center')
    ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
    ALIGN_JUSTIFY = Alignment(horizontal='justify', vertical='top', wrap_text=True)

    # Borders
    BORDER_THIN = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    BORDER_THICK = Border(
        left=Side(style='thick', color='000000'),
        right=Side(style='thick', color='000000'),
        top=Side(style='thick', color='000000'),
        bottom=Side(style='thick', color='000000')
    )

    # Fills
    FILL_HEADER = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    FILL_ACCENT = PatternFill(start_color=COLOR_ACCENT, end_color=COLOR_ACCENT, fill_type='solid')
    FILL_LIGHT = PatternFill(start_color=COLOR_LIGHT, end_color=COLOR_LIGHT, fill_type='solid')


class EnhancedExcelExporter:
    """Export data to professional Excel files with formatting and charts."""

    def __init__(self):
        """Initialize enhanced Excel exporter."""
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for enhanced Excel export. "
                "Install it with: pip install openpyxl"
            )
        self.workbook = None
        self.sheets = {}

    def create_workbook(self, title: str = "Export") -> Workbook:
        """
        Create a new workbook.

        Args:
            title: Workbook title

        Returns:
            Workbook instance
        """
        self.workbook = Workbook()
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']

        # Set workbook properties
        self.workbook.properties.title = title
        self.workbook.properties.creator = "Document Management System"
        self.workbook.properties.created = datetime.now()

        return self.workbook

    def add_sheet(self, name: str, data: List[Dict[str, Any]],
                  headers: Optional[List[str]] = None,
                  title: Optional[str] = None,
                  style_header: bool = True,
                  auto_filter: bool = True,
                  freeze_panes: bool = True) -> Any:
        """
        Add a sheet with data and formatting.

        Args:
            name: Sheet name
            data: List of dictionaries with data
            headers: Optional custom headers (default: use dict keys)
            title: Optional title row
            style_header: Apply header styling
            auto_filter: Add auto-filter to headers
            freeze_panes: Freeze header row

        Returns:
            Worksheet object
        """
        if not self.workbook:
            self.create_workbook()

        # Create sheet
        ws = self.workbook.create_sheet(title=name)
        self.sheets[name] = ws

        # Add title if provided
        start_row = 1
        if title:
            ws.merge_cells(f'A1:{get_column_letter(len(headers or list(data[0].keys())))}1')
            cell = ws['A1']
            cell.value = title
            cell.font = ExcelStyle.FONT_TITLE
            cell.alignment = ExcelStyle.ALIGN_CENTER
            start_row = 3

        # Get headers
        if not headers and data:
            headers = list(data[0].keys())

        # Write headers
        if headers:
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=start_row, column=col_idx)
                cell.value = header

                if style_header:
                    cell.font = ExcelStyle.FONT_HEADER
                    cell.fill = ExcelStyle.FILL_HEADER
                    cell.alignment = ExcelStyle.ALIGN_CENTER
                    cell.border = ExcelStyle.BORDER_THIN

        # Write data
        for row_idx, row_data in enumerate(data, start=start_row + 1):
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = row_data.get(header, '')

                # Format values
                if isinstance(value, (datetime, date)):
                    cell.value = value
                    cell.number_format = 'YYYY-MM-DD HH:MM:SS' if isinstance(value, datetime) else 'YYYY-MM-DD'
                elif isinstance(value, (int, float, Decimal)):
                    cell.value = float(value)
                    cell.number_format = '#,##0.00' if isinstance(value, (float, Decimal)) else '#,##0'
                elif isinstance(value, bool):
                    cell.value = 'Yes' if value else 'No'
                else:
                    cell.value = str(value)

                cell.border = ExcelStyle.BORDER_THIN
                cell.alignment = ExcelStyle.ALIGN_LEFT

        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Add auto-filter
        if auto_filter and headers:
            ws.auto_filter.ref = ws.dimensions

        # Freeze panes
        if freeze_panes:
            ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

        return ws

    def add_chart(self, sheet_name: str, chart_type: str,
                  data_range: str, title: str,
                  position: str = 'E2',
                  width: int = 15, height: int = 10) -> None:
        """
        Add a chart to a sheet.

        Args:
            sheet_name: Sheet name
            chart_type: 'bar', 'line', 'pie', or 'area'
            data_range: Excel range for data (e.g., 'A2:B10')
            title: Chart title
            position: Cell position for chart (e.g., 'E2')
            width: Chart width in columns
            height: Chart height in rows
        """
        if sheet_name not in self.sheets:
            logger.error(f"Sheet '{sheet_name}' not found")
            return

        ws = self.sheets[sheet_name]

        # Create chart based on type
        chart_map = {
            'bar': BarChart,
            'line': LineChart,
            'pie': PieChart,
            'area': AreaChart
        }

        ChartClass = chart_map.get(chart_type.lower(), BarChart)
        chart = ChartClass()
        chart.title = title
        chart.style = 10
        chart.width = width
        chart.height = height

        # Add data to chart
        data = Reference(ws, range_string=data_range)
        chart.add_data(data, titles_from_data=True)

        # Add chart to worksheet
        ws.add_chart(chart, position)

        logger.info(f"Added {chart_type} chart '{title}' to sheet '{sheet_name}'")

    def add_conditional_formatting(self, sheet_name: str,
                                   range_: str,
                                   rule_type: str,
                                   **kwargs) -> None:
        """
        Add conditional formatting to a range.

        Args:
            sheet_name: Sheet name
            range_: Excel range (e.g., 'A2:A100')
            rule_type: 'color_scale', 'cell_is', or 'formula'
            **kwargs: Additional arguments for the rule
        """
        if sheet_name not in self.sheets:
            logger.error(f"Sheet '{sheet_name}' not found")
            return

        ws = self.sheets[sheet_name]

        if rule_type == 'color_scale':
            rule = ColorScaleRule(
                start_type='min', start_color='FF0000',
                mid_type='percentile', mid_value=50, mid_color='FFFF00',
                end_type='max', end_color='00FF00'
            )
        elif rule_type == 'cell_is':
            operator = kwargs.get('operator', 'greaterThan')
            formula = kwargs.get('formula', ['0'])
            fill = PatternFill(start_color=kwargs.get('color', 'FF0000'),
                             end_color=kwargs.get('color', 'FF0000'),
                             fill_type='solid')
            rule = CellIsRule(operator=operator, formula=formula, fill=fill)
        elif rule_type == 'formula':
            formula = kwargs.get('formula', ['TRUE'])
            fill = PatternFill(start_color=kwargs.get('color', 'FF0000'),
                             end_color=kwargs.get('color', 'FF0000'),
                             fill_type='solid')
            rule = FormulaRule(formula=formula, fill=fill)
        else:
            logger.error(f"Unknown rule type: {rule_type}")
            return

        ws.conditional_formatting.add(range_, rule)
        logger.info(f"Added conditional formatting to '{range_}' in sheet '{sheet_name}'")

    def add_data_validation(self, sheet_name: str,
                           range_: str,
                           validation_type: str,
                           **kwargs) -> None:
        """
        Add data validation to a range.

        Args:
            sheet_name: Sheet name
            range_: Excel range (e.g., 'A2:A100')
            validation_type: 'list', 'whole', 'decimal', 'date', etc.
            **kwargs: Additional validation parameters
        """
        if sheet_name not in self.sheets:
            logger.error(f"Sheet '{sheet_name}' not found")
            return

        ws = self.sheets[sheet_name]

        dv = DataValidation(type=validation_type, **kwargs)
        ws.add_data_validation(dv)
        dv.add(range_)

        logger.info(f"Added data validation to '{range_}' in sheet '{sheet_name}'")

    def add_formula(self, sheet_name: str, cell: str, formula: str) -> None:
        """
        Add a formula to a cell.

        Args:
            sheet_name: Sheet name
            cell: Cell address (e.g., 'A1')
            formula: Excel formula (e.g., '=SUM(A2:A10)')
        """
        if sheet_name not in self.sheets:
            logger.error(f"Sheet '{sheet_name}' not found")
            return

        ws = self.sheets[sheet_name]
        ws[cell] = formula
        ws[cell].font = ExcelStyle.FONT_BOLD

        logger.info(f"Added formula '{formula}' to cell '{cell}' in sheet '{sheet_name}'")

    def save(self, output_path: str) -> bool:
        """
        Save workbook to file.

        Args:
            output_path: Output file path

        Returns:
            True if successful
        """
        if not self.workbook:
            logger.error("No workbook to save")
            return False

        try:
            output_path = Path(output_path)
            output_path.parent.mkdir(parents=True, exist_ok=True)

            self.workbook.save(str(output_path))
            logger.info(f"Saved workbook to: {output_path}")
            return True

        except Exception as e:
            logger.error(f"Error saving workbook: {e}")
            return False

    def export_services_report(self, services: List[Dict[str, Any]],
                               output_path: str) -> bool:
        """
        Export services report with multiple sheets and charts.

        Args:
            services: List of service dictionaries
            output_path: Output file path

        Returns:
            True if successful
        """
        try:
            self.create_workbook("Services Report")

            # Sheet 1: Services Overview
            self.add_sheet(
                name="Services Overview",
                data=services,
                title="Services Overview Report",
                style_header=True,
                auto_filter=True,
                freeze_panes=True
            )

            # Sheet 2: Financial Summary
            if services:
                financial_data = [
                    {
                        'Service': s.get('service_name', 'N/A'),
                        'Region': s.get('region', 'N/A'),
                        'Rate': s.get('brutto_rate', 0),
                        'Total Cost': s.get('total_cost', 0)
                    }
                    for s in services
                ]

                ws = self.add_sheet(
                    name="Financial Summary",
                    data=financial_data,
                    title="Financial Summary",
                    style_header=True
                )

                # Add totals row
                last_row = len(financial_data) + 4
                ws[f'C{last_row}'] = 'TOTAL:'
                ws[f'C{last_row}'].font = ExcelStyle.FONT_BOLD
                ws[f'D{last_row}'] = f'=SUM(D4:D{last_row-1})'
                ws[f'D{last_row}'].font = ExcelStyle.FONT_BOLD
                ws[f'D{last_row}'].number_format = '#,##0.00'
                ws[f'E{last_row}'] = f'=SUM(E4:E{last_row-1})'
                ws[f'E{last_row}'].font = ExcelStyle.FONT_BOLD
                ws[f'E{last_row}'].number_format = '#,##0.00'

                # Add conditional formatting
                self.add_conditional_formatting(
                    sheet_name="Financial Summary",
                    range_=f'D4:D{last_row-1}',
                    rule_type='color_scale'
                )

            # Save workbook
            return self.save(output_path)

        except Exception as e:
            logger.error(f"Error exporting services report: {e}")
            return False


# Convenience functions
def export_to_excel(data: List[Dict[str, Any]],
                   output_path: str,
                   sheet_name: str = "Data",
                   title: Optional[str] = None) -> bool:
    """
    Quick export data to Excel.

    Args:
        data: List of dictionaries
        output_path: Output file path
        sheet_name: Sheet name
        title: Optional title

    Returns:
        True if successful
    """
    exporter = EnhancedExcelExporter()
    exporter.create_workbook()
    exporter.add_sheet(sheet_name, data, title=title)
    return exporter.save(output_path)


def export_with_charts(data: List[Dict[str, Any]],
                      output_path: str,
                      chart_config: Optional[Dict[str, Any]] = None) -> bool:
    """
    Export data with charts.

    Args:
        data: List of dictionaries
        output_path: Output file path
        chart_config: Chart configuration

    Returns:
        True if successful
    """
    exporter = EnhancedExcelExporter()
    exporter.create_workbook()

    sheet_name = "Data"
    exporter.add_sheet(sheet_name, data)

    if chart_config:
        exporter.add_chart(
            sheet_name=sheet_name,
            chart_type=chart_config.get('type', 'bar'),
            data_range=chart_config.get('range', 'A1:B10'),
            title=chart_config.get('title', 'Chart'),
            position=chart_config.get('position', 'E2')
        )

    return exporter.save(output_path)
